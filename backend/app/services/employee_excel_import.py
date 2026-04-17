"""Парсинг Excel со списком сотрудников (учётка, ФИО, должность, подразделение, системы)."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass

from openpyxl import load_workbook


def _norm_header(val: object) -> str:
    if val is None:
        return ""
    s = str(val).strip().lower()
    return re.sub(r"\s+", "", s)


_LOGIN_HEADERS = frozenset({"учетнаязапись", "логин", "login"})
_NAME_HEADERS = frozenset({"фио", "фио.", "ф.и.о."})
_POSITION_HEADERS = frozenset(
    {
        "должность",
        "должностьсотрудника",
        "позиция",
        "jobtitle",
        "title",
    }
)
_DEPARTMENT_HEADERS = frozenset({"подразделение"})
_SYSTEMS_HEADERS = frozenset({"системы", "система"})


@dataclass(frozen=True)
class ParsedEmployeeRow:
    sheet_row: int
    login: str
    full_name: str
    position_title: str | None
    department_title: str | None
    systems: tuple[str, ...]


def parse_employee_excel_xlsx(content: bytes) -> tuple[list[ParsedEmployeeRow], str | None]:
    """
    Читает первый лист .xlsx, ищет строку заголовков с колонками учётной записи, ФИО и должности.
    Дополнительно (если есть): «Подразделение», «Системы».
    Возвращает (строки данных, сообщение об ошибке парсинга или None).
    """
    try:
        wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    except Exception as e:  # noqa: BLE001
        return [], f"Не удалось открыть файл как Excel (.xlsx): {e!s}"

    try:
        if not wb.sheetnames:
            return [], "В файле нет листов"

        ws = wb[wb.sheetnames[0]]
        header_map: dict[str, int] | None = None
        header_row_idx = 0

        max_scan = min(ws.max_row or 0, 50)
        for r in range(1, max_scan + 1):
            row_vals: list[object] = []
            for c in range(1, (ws.max_column or 0) + 1):
                row_vals.append(ws.cell(r, c).value)
            mapping: dict[str, int] = {}
            for col_idx, raw in enumerate(row_vals, start=1):
                key = _norm_header(raw)
                if not key:
                    continue
                if key in _LOGIN_HEADERS:
                    mapping["login"] = col_idx
                elif key in _NAME_HEADERS:
                    mapping["full_name"] = col_idx
                elif key in _POSITION_HEADERS:
                    mapping["position"] = col_idx
                elif key in _DEPARTMENT_HEADERS:
                    mapping["department"] = col_idx
                elif key in _SYSTEMS_HEADERS:
                    mapping["systems"] = col_idx
            if "login" in mapping and "full_name" in mapping and "position" in mapping:
                header_map = mapping
                header_row_idx = r
                break

        if not header_map:
            return [], "Не найдена строка заголовков с колонками «УчетнаяЗапись», «ФИО» и «Должность»"

        col_login = header_map["login"]
        col_name = header_map["full_name"]
        col_pos = header_map["position"]
        col_dep = header_map.get("department")
        col_sys = header_map.get("systems")

        out: list[ParsedEmployeeRow] = []
        for r in range(header_row_idx + 1, (ws.max_row or 0) + 1):
            raw_login = ws.cell(r, col_login).value
            raw_name = ws.cell(r, col_name).value
            raw_pos = ws.cell(r, col_pos).value
            raw_dep = ws.cell(r, col_dep).value if col_dep else None
            raw_sys = ws.cell(r, col_sys).value if col_sys else None

            login = _cell_str(raw_login)
            full_name = _cell_str(raw_name)
            pos_t = _cell_str(raw_pos) if raw_pos is not None and str(raw_pos).strip() else None
            dep_t = _cell_str(raw_dep) if raw_dep is not None and str(raw_dep).strip() else None
            systems = _parse_systems_cell(raw_sys)

            if not login and not full_name:
                continue
            out.append(
                ParsedEmployeeRow(
                    sheet_row=r,
                    login=login,
                    full_name=full_name,
                    position_title=pos_t,
                    department_title=dep_t,
                    systems=systems,
                )
            )

        return out, None
    finally:
        wb.close()


def _cell_str(val: object) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _parse_systems_cell(val: object) -> tuple[str, ...]:
    s = _cell_str(val)
    if not s:
        return ()
    parts = re.split(r"[;,/\n]+", s)
    out: list[str] = []
    seen: set[str] = set()
    for p in parts:
        name = " ".join(p.split())
        if not name:
            continue
        k = name.casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(name)
    return tuple(out)
