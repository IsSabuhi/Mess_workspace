"""Импорт расписания месяца из Excel-листа."""

from __future__ import annotations

import calendar
import io
import uuid

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import ScheduleEntry, User

_MONTH_SHEET_NAMES = {
    1: "январь",
    2: "февраль",
    3: "март",
    4: "апрель",
    5: "май",
    6: "июнь",
    7: "июль",
    8: "август",
    9: "сентябрь",
    10: "октябрь",
    11: "ноябрь",
    12: "декабрь",
}


def _norm_name(s: str) -> str:
    return " ".join((s or "").split()).strip().lower()


def _cell_str(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _pick_sheet(wb, month: int, sheet_name: str | None):
    if sheet_name:
        for n in wb.sheetnames:
            if _norm_name(n) == _norm_name(sheet_name):
                return wb[n]
        return None

    target = _MONTH_SHEET_NAMES.get(month, "")
    for n in wb.sheetnames:
        if _norm_name(n) == target:
            return wb[n]
    return wb[wb.sheetnames[0]] if wb.sheetnames else None


def _discover_day_columns(ws, dim: int) -> tuple[int, dict[int, int]]:
    """
    Возвращает (row_idx, {day: col_idx}) для строки с номерами дней.
    Ищем строку с максимальным числом ячеек-чисел 1..dim.
    """
    best_row = 0
    best_map: dict[int, int] = {}
    max_scan = min(ws.max_row or 0, 30)
    max_col = ws.max_column or 0
    for r in range(1, max_scan + 1):
        m: dict[int, int] = {}
        for c in range(1, max_col + 1):
            raw = ws.cell(r, c).value
            s = _cell_str(raw)
            if not s.isdigit():
                continue
            d = int(s)
            if 1 <= d <= dim and d not in m:
                m[d] = c
        if len(m) > len(best_map):
            best_map = m
            best_row = r
    return best_row, best_map


def parse_month_schedule_sheet(content: bytes, *, month: int, dim: int, sheet_name: str | None) -> tuple[dict[str, dict[int, str]], str | None]:
    """Парсит лист месяца: {normalized_full_name: {day: code}}."""
    try:
        wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    except Exception as e:  # noqa: BLE001
        return {}, f"Не удалось открыть Excel: {e!s}"

    try:
        ws = _pick_sheet(wb, month, sheet_name)
        if ws is None:
            return {}, "В файле нет листов"

        day_row, day_cols = _discover_day_columns(ws, dim)
        if not day_row or len(day_cols) < max(10, dim // 2):
            return {}, "Не удалось определить строку с днями месяца"

        first_day_col = min(day_cols.values())
        name_col = max(1, first_day_col - 1)

        out: dict[str, dict[int, str]] = {}
        for r in range(day_row + 1, (ws.max_row or 0) + 1):
            name = _cell_str(ws.cell(r, name_col).value)
            if not name:
                continue
            nk = _norm_name(name)
            if not nk:
                continue
            row_map: dict[int, str] = {}
            has_any = False
            for d in range(1, dim + 1):
                c = day_cols.get(d)
                if not c:
                    continue
                code = _cell_str(ws.cell(r, c).value)
                if code:
                    has_any = True
                row_map[d] = code
            if has_any:
                out[nk] = row_map
        return out, None
    finally:
        wb.close()


async def import_schedule_month_excel(
    session: AsyncSession,
    *,
    year: int,
    month: int,
    content: bytes,
    editor_id: uuid.UUID,
    sheet_name: str | None = None,
) -> dict:
    dim = calendar.monthrange(year, month)[1]
    parsed, err = parse_month_schedule_sheet(content, month=month, dim=dim, sheet_name=sheet_name)
    if err:
        return {"error": err}

    users = (await session.execute(select(User).where(User.is_active.is_(True)))).scalars().all()
    by_name: dict[str, User] = {}
    for u in users:
        nk = _norm_name(u.full_name)
        if nk and nk not in by_name:
            by_name[nk] = u

    matched: list[tuple[User, dict[int, str]]] = []
    unmatched_names: list[str] = []
    for nk, day_map in parsed.items():
        u = by_name.get(nk)
        if not u:
            unmatched_names.append(nk)
            continue
        matched.append((u, day_map))

    if matched:
        user_ids = [u.id for u, _ in matched]
        await session.execute(
            delete(ScheduleEntry).where(
                ScheduleEntry.year == year,
                ScheduleEntry.month == month,
                ScheduleEntry.user_id.in_(user_ids),
            )
        )

    inserted = 0
    for u, row_map in matched:
        for d in range(1, dim + 1):
            code = (row_map.get(d) or "").strip()
            if not code:
                continue
            session.add(
                ScheduleEntry(
                    year=year,
                    month=month,
                    day=d,
                    user_id=u.id,
                    code=code,
                    updated_by_id=editor_id,
                )
            )
            inserted += 1

    await session.commit()
    return {
        "year": year,
        "month": month,
        "sheet_used": sheet_name or _MONTH_SHEET_NAMES.get(month, ""),
        "users_matched": len(matched),
        "rows_parsed": len(parsed),
        "cells_imported": inserted,
        "unmatched_names": sorted(unmatched_names),
    }
