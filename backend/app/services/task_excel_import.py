"""Импорт задач из Excel-задачника (листы вида «СМЗиС»)."""

from __future__ import annotations

import io
import re
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Board, KanbanColumn, System, Task, User
from app.models.task import TaskPriority
from app.schemas.task_excel_import import (
    TaskExcelImportBatchOut,
    TaskExcelImportFileResult,
    TaskExcelImportOut,
    TaskExcelImportRowDetail,
)

_HEADER_ALIASES = {
    "num": {"№ п/п", "№", "n", "num", "номер"},
    "title": {"задача", "название", "title"},
    "assignees": {"исполнитель", "исполнители", "assignee", "assignees"},
    "mark": {"!", "статус", "mark"},
    "start": {"начало", "start", "дата начала"},
    "end": {"окончание", "срок", "end", "due"},
    "closed": {"дата закрытия", "закрытие", "closed"},
    "comment": {"комментарий", "comment", "описание"},
}


@dataclass
class ParsedTaskRow:
    row_num: int
    title: str
    assignee_tokens: list[str] = field(default_factory=list)
    mark: str = ""
    start: datetime | None = None
    end: datetime | None = None
    closed_at: datetime | None = None
    comment: str = ""


def _norm(s: str) -> str:
    return " ".join((s or "").split()).strip().lower()


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    if isinstance(v, datetime):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, date):
        return v.strftime("%d.%m.%Y")
    return str(v).strip()


def _as_datetime(v: Any) -> datetime | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        if v.tzinfo is None:
            return v.replace(tzinfo=timezone.utc)
        return v.astimezone(timezone.utc)
    if isinstance(v, date):
        return datetime.combine(v, time.min, tzinfo=timezone.utc)
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _pick_data_sheet(wb, preferred: str | None) -> Worksheet | None:
    names = [n for n in wb.sheetnames if "cognos" not in _norm(n)]
    if not names:
        return None
    if preferred:
        pref = _norm(preferred)
        for n in names:
            if _norm(n) == pref:
                return wb[n]
        for n in names:
            if pref in _norm(n) or _norm(n) in pref:
                return wb[n]
    # первый «нормальный» лист
    return wb[names[0]]


def _find_header(ws: Worksheet) -> tuple[int, dict[str, int]] | None:
    max_r = min(ws.max_row or 0, 40)
    max_c = min(ws.max_column or 0, 30)
    for r in range(1, max_r + 1):
        mapping: dict[str, int] = {}
        for c in range(1, max_c + 1):
            raw = _cell_str(ws.cell(r, c).value)
            if not raw:
                continue
            key = _norm(raw).rstrip(":")
            for field_name, aliases in _HEADER_ALIASES.items():
                if key in aliases or key == field_name:
                    mapping[field_name] = c
                    break
            if key == "!" or raw == "!":
                mapping["mark"] = c
        if "title" in mapping:
            return r, mapping
    return None


def _split_assignees(raw: str) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[,;/]| и ", raw)
    out: list[str] = []
    for p in parts:
        t = _norm(p)
        if t:
            # берём первое слово — фамилию
            out.append(t.split()[0])
    return out


def parse_task_workbook(
    content: bytes, *, sheet_name: str | None = None
) -> tuple[list[ParsedTaskRow], str | None, str | None]:
    """Возвращает (rows, sheet_name, error)."""
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:  # noqa: BLE001
        return [], None, f"Не удалось открыть Excel: {e!s}"

    ws = _pick_data_sheet(wb, sheet_name)
    if ws is None:
        return [], None, "В файле нет подходящих листов"

    found = _find_header(ws)
    if not found:
        return [], ws.title, "Не найдена строка заголовков (ожидается колонка «Задача»)"

    header_row, cols = found
    title_c = cols["title"]
    rows: list[ParsedTaskRow] = []

    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        title = _cell_str(ws.cell(r, title_c).value)
        if not title:
            continue
        if len(title) > 512:
            title = title[:509] + "…"

        mark = _cell_str(ws.cell(r, cols["mark"]).value) if "mark" in cols else ""
        assignees_raw = _cell_str(ws.cell(r, cols["assignees"]).value) if "assignees" in cols else ""
        start = _as_datetime(ws.cell(r, cols["start"]).value) if "start" in cols else None
        end = _as_datetime(ws.cell(r, cols["end"]).value) if "end" in cols else None
        closed = _as_datetime(ws.cell(r, cols["closed"]).value) if "closed" in cols else None
        comment = _cell_str(ws.cell(r, cols["comment"]).value) if "comment" in cols else ""

        rows.append(
            ParsedTaskRow(
                row_num=r,
                title=title,
                assignee_tokens=_split_assignees(assignees_raw),
                mark=mark,
                start=start,
                end=end,
                closed_at=closed,
                comment=comment,
            )
        )

    if not rows:
        return [], ws.title, "На листе нет строк с задачами"

    return rows, ws.title, None


def _is_closed(row: ParsedTaskRow) -> bool:
    if row.closed_at is not None:
        return True
    m = _norm(row.mark)
    return m in {"v", "✓", "да", "done", "выполнено", "закрыто"}


def _priority(row: ParsedTaskRow) -> TaskPriority:
    m = _norm(row.mark)
    if m in {"!", "1", "urgent", "срочно", "важно"}:
        return TaskPriority.high
    return TaskPriority.normal


def _build_description(row: ParsedTaskRow, sheet_name: str) -> str | None:
    parts: list[str] = []
    if row.comment:
        parts.append(row.comment)
    meta: list[str] = []
    if row.start:
        meta.append(f"Начало: {row.start.strftime('%d.%m.%Y')}")
    if row.end:
        meta.append(f"Окончание: {row.end.strftime('%d.%m.%Y')}")
    if row.closed_at:
        meta.append(f"Дата закрытия: {row.closed_at.strftime('%d.%m.%Y')}")
    if meta:
        parts.append(" · ".join(meta))
    parts.append(f"Импорт из задачника Excel (лист «{sheet_name}»)")
    text = "\n".join(parts).strip()
    return text or None


def _surname(full_name: str) -> str:
    parts = _norm(full_name).split()
    return parts[0] if parts else ""


async def import_tasks_from_excel(
    session: AsyncSession,
    *,
    content: bytes,
    system_id: uuid.UUID,
    creator_id: uuid.UUID,
    sheet_name: str | None = None,
) -> TaskExcelImportOut | dict[str, str]:
    system = await session.get(System, system_id)
    if not system or not system.is_active:
        return {"error": "Система не найдена или неактивна"}

    board = await session.scalar(select(Board).where(Board.is_default.is_(True)))
    if not board:
        return {"error": "Основная доска не найдена"}

    columns = list(
        (
            await session.execute(
                select(KanbanColumn)
                .where(KanbanColumn.board_id == board.id)
                .order_by(KanbanColumn.sort_order)
            )
        )
        .scalars()
        .all()
    )
    if not columns:
        return {"error": "У основной доски нет колонок"}

    done_col = next((c for c in columns if c.is_done_column), None)
    in_progress = next(
        (c for c in columns if c.slug == "in_progress" or "работ" in (c.name or "").lower()),
        None,
    )
    if in_progress is None:
        in_progress = next((c for c in columns if not c.is_done_column), None)
    if in_progress is None:
        return {"error": "Не найдена колонка «В работе»"}
    if done_col is None:
        return {"error": "Не найдена колонка «Выполнено» (is_done_column)"}

    parsed, used_sheet, err = parse_task_workbook(content, sheet_name=sheet_name)
    if err:
        return {"error": err}
    assert used_sheet is not None

    users = list(
        (await session.execute(select(User).where(User.is_active.is_(True)))).scalars().unique().all()
    )
    by_surname: dict[str, list[User]] = {}
    for u in users:
        sn = _surname(u.full_name)
        if sn:
            by_surname.setdefault(sn, []).append(u)

    # позиции в колонках
    pos_done = int(
        await session.scalar(
            select(func.coalesce(func.max(Task.position), -1)).where(Task.column_id == done_col.id)
        )
        or -1
    )
    pos_ip = int(
        await session.scalar(
            select(func.coalesce(func.max(Task.position), -1)).where(Task.column_id == in_progress.id)
        )
        or -1
    )

    details: list[TaskExcelImportRowDetail] = []
    created = skipped = warnings = 0

    for row in parsed:
        closed = _is_closed(row)
        column = done_col if closed else in_progress
        if closed:
            pos_done += 1
            position = pos_done
        else:
            pos_ip += 1
            position = pos_ip

        matched: list[User] = []
        warn_msgs: list[str] = []
        for token in row.assignee_tokens:
            cands = by_surname.get(token, [])
            if len(cands) == 1:
                matched.append(cands[0])
            elif len(cands) == 0:
                warn_msgs.append(f"не найден исполнитель «{token}»")
            else:
                warn_msgs.append(f"несколько пользователей с фамилией «{token}» — пропуск")

        # уникальные
        seen: set[uuid.UUID] = set()
        assignees: list[User] = []
        for u in matched:
            if u.id not in seen:
                seen.add(u.id)
                assignees.append(u)

        task = Task(
            title=row.title,
            description=_build_description(row, used_sheet),
            board_id=board.id,
            column_id=column.id,
            system_id=system.id,
            creator_id=creator_id,
            priority=_priority(row),
            due_at=row.end,
            checklist=[],
            position=position,
        )
        if row.start is not None:
            task.created_at = row.start
        task.assignees = assignees
        session.add(task)
        await session.flush()

        created += 1
        status = "created"
        msg = None
        if warn_msgs:
            warnings += 1
            status = "warning"
            msg = "; ".join(warn_msgs)
            if closed:
                msg = f"В «Выполнено». {msg}"
            else:
                msg = f"В «В работе». {msg}"
        else:
            msg = "В «Выполнено»" if closed else "В «В работе»"

        details.append(
            TaskExcelImportRowDetail(
                row=row.row_num,
                title=row.title,
                status=status,
                message=msg,
                task_id=task.id,
            )
        )

    await session.flush()
    return TaskExcelImportOut(
        sheet_name=used_sheet,
        system_id=system.id,
        system_name=system.name,
        board_id=board.id,
        created=created,
        skipped=skipped,
        warnings=warnings,
        rows=details,
    )


def _filename_token(filename: str) -> str:
    stem = filename.rsplit("/", 1)[-1]
    if "." in stem:
        stem = stem.rsplit(".", 1)[0]
    t = _norm(stem)
    t = re.sub(r"^задачник\s+", "", t)
    t = t.replace("_", " ").replace("-", " ")
    return " ".join(t.split())


def match_system_for_import(systems: list[System], *, filename: str, sheet_name: str) -> System | None:
    """Сопоставить файл/лист с системой по имени или slug."""
    tokens: list[str] = []
    for raw in (sheet_name, _filename_token(filename)):
        t = _norm(raw)
        if t and t not in tokens:
            tokens.append(t)
    if not tokens:
        return None

    best: System | None = None
    best_score = 0
    for sys in systems:
        name_n = _norm(sys.name)
        slug_n = _norm(sys.slug or "")
        name_short = re.sub(r"\([^)]*\)", "", name_n).strip()
        for tok in tokens:
            score = 0
            if tok and tok in {slug_n, name_n, name_short}:
                score = 100
            elif tok and (tok in name_n or (name_short and (name_short in tok or tok in name_short))):
                score = 80
            elif tok and slug_n and (tok in slug_n or slug_n in tok):
                score = 70
            if score > best_score:
                best_score = score
                best = sys
    return best if best_score >= 70 else None


async def import_tasks_from_excel_batch(
    session: AsyncSession,
    *,
    files: list[tuple[str, bytes]],
    creator_id: uuid.UUID,
    system_id: uuid.UUID | None = None,
    sheet_name: str | None = None,
) -> TaskExcelImportBatchOut:
    """Несколько файлов: система из system_id (если один файл) или авто-матч по имени/листу."""
    systems = list(
        (await session.execute(select(System).where(System.is_active.is_(True)).order_by(System.name)))
        .scalars()
        .all()
    )
    out_files: list[TaskExcelImportFileResult] = []
    created_total = 0
    warnings_total = 0
    ok_n = 0
    fail_n = 0

    for filename, content in files:
        forced_system_id = system_id if (system_id is not None and len(files) == 1) else None
        resolved_id = forced_system_id

        if resolved_id is None:
            # Нужен sheet для матча — лёгкий parse
            _rows, used_sheet, parse_err = parse_task_workbook(content, sheet_name=sheet_name)
            if parse_err:
                fail_n += 1
                out_files.append(
                    TaskExcelImportFileResult(filename=filename, ok=False, error=parse_err)
                )
                continue
            matched = match_system_for_import(systems, filename=filename, sheet_name=used_sheet or "")
            if matched is None:
                fail_n += 1
                out_files.append(
                    TaskExcelImportFileResult(
                        filename=filename,
                        ok=False,
                        error=(
                            f"Не удалось сопоставить систему (лист «{used_sheet}»). "
                            "Переименуйте файл/лист ближе к названию системы или загрузите один файл и укажите систему вручную."
                        ),
                    )
                )
                continue
            resolved_id = matched.id

        res = await import_tasks_from_excel(
            session,
            content=content,
            system_id=resolved_id,
            creator_id=creator_id,
            sheet_name=sheet_name,
        )
        if isinstance(res, dict) and "error" in res:
            fail_n += 1
            out_files.append(
                TaskExcelImportFileResult(filename=filename, ok=False, error=str(res["error"]))
            )
            continue

        assert isinstance(res, TaskExcelImportOut)
        ok_n += 1
        created_total += res.created
        warnings_total += res.warnings
        out_files.append(TaskExcelImportFileResult(filename=filename, ok=True, result=res))

    await session.flush()
    return TaskExcelImportBatchOut(
        files_total=len(files),
        files_ok=ok_n,
        files_failed=fail_n,
        created_total=created_total,
        warnings_total=warnings_total,
        files=out_files,
    )
